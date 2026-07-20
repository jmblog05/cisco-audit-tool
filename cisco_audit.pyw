import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
from tkinter import messagebox
import os
import re
import glob
from openpyxl import Workbook
from openpyxl import load_workbook  
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment # for wrap text
import ipaddress

subnet_mask_to_prefix = {
    '255.255.255.255': 32,
    '255.255.255.254': 31,
    '255.255.255.252': 30,
    '255.255.255.248': 29,
    '255.255.255.240': 28,
    '255.255.255.224': 27,
    '255.255.255.192': 26,
    '255.255.255.128': 25,
    '255.255.255.0': 24,
    '255.255.254.0': 23,
    '255.255.252.0': 22,
    '255.255.248.0': 21,
    '255.255.240.0': 20,
    '255.255.224.0': 19,
    '255.255.192.0': 18,
    '255.255.128.0': 17,
    '255.255.0.0': 16,
    '255.254.0.0': 15,
    '255.252.0.0': 14,
    '255.248.0.0': 13,
    '255.240.0.0': 12,
    '255.224.0.0': 11,
    '255.192.0.0': 10,
    '255.128.0.0': 9,
    '255.0.0.0': 8,
}

def initial_screen(): 
    global log_directory     
    global check1
    global check2    
    root = tk.Tk()
    root.title("Cisco Audit tool")
    root.geometry("1000x215")

    
    canvas = tk.Canvas(root)

    bar = tk.Scrollbar(root, orient=tk.VERTICAL)
    bar.pack(side=tk.RIGHT, fill=tk.Y)
    bar.config(command=canvas.yview) 

    canvas.config(yscrollcommand=bar.set) 
    canvas.config(scrollregion=(0,0,1000,215)) 
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    frame = tk.Frame(canvas)
    canvas.create_window((0,0), window=frame, anchor=tk.NW, width=1000, height=215)
   
    log_directory = tk.StringVar(value="")
    check1 = tk.IntVar(value=1)
    check2 = tk.IntVar(value=1)
    font_general = ('', '14', 'normal')

    s = ttk.Style()
    s.theme_use('default') #'winnative', 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative')
    s.configure('MyWidget.TButton', background='royalblue', foreground='lightgreen',font = (u'', 14,"bold"))
    s.configure('MyWidget2.TButton', background='skyblue', foreground='purple',font = (u'', 14,"bold"))

    label1 = tk.Label(frame, justify="center",text='Step1: Get the following command outputs by CLI and then save the log file into a directory.             ',foreground="olive",font= ('',14,"bold")) 
    label1.grid(row=0, column=0,sticky=tk.W)

    button_1 = ttk.Button(frame, style='MyWidget.TButton', text="Show the command list",command=lambda:[cmd_list()])
    button_1.grid(row=1, column=0,sticky=tk.W)  

    label1 = tk.Label(frame, justify="center",text='Step2: Select the log file directory.             ',foreground="olive",font= ('',14,"bold")) 
    label1.grid(row=2, column=0,sticky=tk.W)

    button_1 = ttk.Button(frame, style='MyWidget.TButton', text="Select the log directory",command=lambda:[select_directory()])
    button_1.grid(row=3, column=0,sticky=tk.W)  

    result_label = tk.Label(frame, justify="center",textvariable=log_directory,foreground="red",font= ('red',14,"")) 
    result_label.place(x =225, y=95)
    
    check_button1 = tk.Checkbutton(frame, text="IP check", variable=check1,font= ('green',12,""))
    check_button1.grid(row=4, column=0,sticky=tk.W)  
    check_button2 = tk.Checkbutton(frame, text="Status check", variable=check2,font= ('gree',12,""))
    check_button2.grid(row=5, column=0,sticky=tk.W)  
    

    button_1 = ttk.Button(frame, style='MyWidget2.TButton', text="Help",command=lambda:[help()])
    button_1.grid(row=6, column=0,sticky=tk.W)  

    
    #frame.pack()
    root.mainloop()

def cmd_list():
    with open(current_directory + r'\tmp.txt', 'w', encoding='utf-8') as f:
        f.write("term len 0\n")
        f.write("show version\n")
        f.write("show run\n")
        f.write("show ip int brief\n")
        f.write("show cdp nei\n")
        f.write("show ip ospf nei\n")
        f.write("show ip bgp summary\n")
        f.write("show ip route\n")
        f.write("show ntp associations\n")
    f.close
    os.popen(current_directory + r'\tmp.txt')

def help():
    with open(current_directory + r'\tmp.txt', 'w', encoding='utf-8') as f:
        f.write("This script analyzes multiple log files containing command outputs collected from Cisco routers and generates the following Excel reports:\n\n")
        f.write("1. IP Address Information for Each Interface\n")
        f.write("   - Interface\n")
        f.write("   - IP Address\n")
        f.write("   - Subnet Mask\n")
        f.write("   - Prefix Length\n")
        f.write("   - Network Address\n")
        f.write("   - Host Range\n")
        f.write("   - Status\n")
        f.write("   - Description\n\n")

        f.write("2. Device Status Information\n")
        f.write("   - Hardware Model\n")
        f.write("   - Software Version\n")
        f.write("   - Uptime\n")
        f.write("   - Interface Status\n")
        f.write("   - CDP Status\n")
        f.write("   - OSPF Neighbor Status\n")
        f.write("   - BGP Neighbor Status\n")
        f.write("   - Default Route Status\n")
        f.write("   - NTP Synchronization Status\n")
    f.close
    os.popen(current_directory + r'\tmp.txt')




def select_directory():
    global log_directory
    root = tk.Tk()
    root.withdraw()
    
    if check1.get() == 0 and check2.get() == 0:
        tk.messagebox.showerror(title="Error", message="Check at least one check box")
    else:
        directory = tk.filedialog.askdirectory()#(initialdir = current_directory)
        file_list = glob.glob(directory +  '\*.log')
        file_list2 = glob.glob(directory +  '\*.txt')
        file_list = file_list + file_list2
        file_list_length = len(file_list)
        log_directory.set(directory)

        try:
            result_file = current_directory + r'\result.xlsx'
            wb = Workbook()
            wb.save(result_file)
        except PermissionError:
            tk.messagebox.showerror(title="Error", message="Close 'result.xlsx'")  
        if file_list_length == 0:  
            log_directory_selected = "no"
        else:
            log_directory_selected = "yes"
        wb = load_workbook(result_file)  
###########################
#IP info
###########################
        if check1.get() == 1:
            ws = wb.active
            ws.title = "IP"
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 8   
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 28
            ws.column_dimensions['H'].width = 10   
            ws.column_dimensions['I'].width = 30
            fill = PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')
            red_fill = PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
            green_fill = PatternFill(patternType='solid',fgColor='00FF00', bgColor='00FF00')
            column_list = ["A","B","C","D","E","F","G","H","I"]
            for g in range(len(column_list)):
                ws[column_list[g] + '1'].fill = fill
                ws[column_list[g] + '1'].font = Font(bold=True)
            ws.auto_filter.ref = "A1:I1"    #Excel's filter
            row_number = 1
            title_row = ["hostname","Interface","IP address","Subnet Mask","Prefix","Network address","Host Range","Status","Description"]
            for s in range(0, 9):
                ws.cell(row=row_number, column=s+1, value=title_row[s])

            if log_directory_selected != "yes":  
                tk.messagebox.showerror(title="Error", message="Select the log file directory")
                log_directory_selected = "yes" #Without this, if "Status" is also checked and not selected log direcgtory, error popup twice 
            else:
                try:
                    for n in range(file_list_length):
                        with open(file_list[n], 'r', encoding='utf-8') as g:
                            lines = ' '.join(g.readlines())   
                            lines = lines.split(' ')
                            lines = [a for a in lines if a != '']   #delete spaces
                            file_name_base = os.path.basename(file_list[n])
                            file_name_base = file_name_base.replace('.txt','')
                            file_name_base = file_name_base.replace('.log','')

                            hostname_indexes = [x for x, y in enumerate(lines) if y == "hostname"]
                            if len(hostname_indexes) != 0:
                                hostname = lines[hostname_indexes[0]+1]
                            else:
                                hostname = file_name_base                    

                            indexes = [x for x, y in enumerate(lines) if y == "interface"]   
                            for m in range(len(indexes)):  
                                interface = ""#initialize                    
                                ip_address = ""#initialize
                                subnet_mask = ""#initialize
                                description = ""#initialize
                                description_list = []#initialize                   
                                if re.match("Loopback" + r'\d+', lines[indexes[m]+1].strip('\n')) or re.match("FastEthernet" + r'\d+/\d+', lines[indexes[m]+1].strip('\n')) or re.match("GigabitEthernet" + r'\d+/\d+', lines[indexes[m]+1].strip('\n')):
                                    interface = lines[indexes[m]+1]

#############################Interface & IP Address & Subnet Mask
                                    for k in range(0,10):
                                        if lines[indexes[m]+1+k] == "interface":#next interface
                                            break
                                        elif lines[indexes[m]+1+k] == "ip" and re.match(r'\d+.\d+.\d+.\d+',lines[indexes[m]+1+k+2]):
                                            ip_address = lines[indexes[m]+1+k+2]                               
                                            subnet_mask = lines[indexes[m]+1+k+3]                                    
                                        elif lines[indexes[m]+1+k] == "description":   
                                            for p in range(1,10):
                                                if lines[indexes[m]+1+k+p] == "ip" and lines[indexes[m]+1+k+p+1] == "address":
                                                    break
                                                else:
                                                    description_list.append(lines[indexes[m]+1+k+p])
                                            description = ' '.join(description_list)                                  
#############################Status                              
                                status = ""#initialize
                                indexes2 = [z for z, w in enumerate(lines) if w == interface.strip('\n')]
                                red_flag = "off"
                                for l in range(len(indexes2)):                                
                                    if lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "manual" or lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "NVRAM" or lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "unset":
                                        if lines[indexes2[l]+5] == "up":
                                            status = "up"
                                            red_flag ="off"
                                        elif lines[indexes2[l]+4] == "administratively":  
                                            status = "admin down"
                                            red_flag ="on"
                                        elif lines[indexes2[l]+4] == "down" and lines[indexes[m]+5] == "down":  
                                            status = "down"  
                                            red_flag ="on"  
                                     
#############################Prefix, network_address, host_range
                                prefix = ""#initialize
                                network_address = ""#initialize
                                host_range = ""#initialize
                                if ip_address != "":                                                 
                                    if subnet_mask.strip('\n') in subnet_mask_to_prefix.keys():
                                        prefix = subnet_mask_to_prefix[subnet_mask.strip('\n')]                     
                                    address_prefix = ipaddress.IPv4Interface(ip_address + "/" + str(prefix))
                                    network_address = str(address_prefix.network)
                                    host_list = list(ipaddress.ip_network(network_address).hosts())  
                                    first_host = str(host_list[0])
                                    last_host = str(host_list[-1])
                                    host_range = (first_host + " to " + last_host)

                                value_list = [hostname,interface,ip_address,subnet_mask,prefix,network_address,host_range,status,description]                            
                                if  ip_address != "":  #ignore interfaces without IP address
                                    row_number +=1
                                    for j in range(len(value_list)):
                                        ws.cell(row=row_number, column=j+1, value=value_list[j])
                                        if red_flag == "on":
                                            ws['H'+str(row_number)].font = Font(color='FFFFFF') #white characters
                                            ws['H'+str(row_number)].fill = red_fill  
                                        else:
                                            ws['H'+str(row_number)].fill = green_fill  

                            side1 = Side(style='thin', color='000000')
                            border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
 
                            for row in ws['A1:I'+str(row_number)]:
                                for cell in row:
                                    cell.border = border_aro #line
                            wb.save(result_file)                                               
                except IndexError:
                    pass
                    
        else:
            wb.remove(wb.worksheets[0]) #new sheet is not created for "IP" so "Sheet" sheet need to be deleted if checkbox for "IP" is not checked.
                    
###########################
#Status info
###########################
        if check2.get() == 1:
            ws = wb.create_sheet('Status') 
            ws = wb['Status']
            wb.active = ws
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 7
            ws.column_dimensions['E'].width = 30 
            ws.column_dimensions['F'].width = 55
            ws.column_dimensions['G'].width = 13
            ws.column_dimensions['H'].width = 12   
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20            
            blue_fill = PatternFill(patternType='solid',fgColor='00FFFF', bgColor='00FF00')
            red_fill = PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
            yellow_fill = PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')
            gray_fill = PatternFill(patternType='solid',fgColor='808080', bgColor='808080')
            green_fill = PatternFill(patternType='solid',fgColor='00FF00', bgColor='00FF00')
            column_list = ["A","B","C","D","E","F","G","H","I","J","K"]
            for g in range(len(column_list)):
                ws[column_list[g] + '1'].fill = yellow_fill
                ws[column_list[g] + '1'].font = Font(bold=True)
            ws.auto_filter.ref = "A1:K1" #Excel's filter       
            row_number2 = 1
            title_row = ["hostname","category1","category2","flag","command","value","","","","",""]
            for s in range(0, 11):
                ws.cell(row=row_number2, column=s+1, value=title_row[s])

            if log_directory_selected != "yes":  
                tk.messagebox.showerror(title="Error", message="Select the log file directory")
            else:
                try:
                    for n in range(file_list_length):
                        with open(file_list[n], 'r', encoding='utf-8') as g:
                            lines = ' '.join(g.readlines())   
                            lines = lines.split(' ')
                            lines = [a for a in lines if a != '']   #delete spaces 
                            file_name_base = os.path.basename(file_list[n])
                            file_name_base = file_name_base.replace('.txt','')
                            file_name_base = file_name_base.replace('.log','')

                            hostname_indexes = [x for x, y in enumerate(lines) if y == "hostname"]
                            if len(hostname_indexes) != 0:
                                hostname = lines[hostname_indexes[0]+1]
                            else:
                                hostname = file_name_base                          

#####
#####show ver
#####
                            hw_model_list = ["7206VXR"]
                            hw_model = "" #initialize
                            sw_ver = "" #initialize
                            uptime_list = [] #initialize
                            uptime = "" #initialize
                            sh_ver_count = 0
                            
                            indexes = [x for x, y in enumerate(lines) if y == "cisco"]   
                            for m in range(len(indexes)): 
                                if lines[indexes[m]+1] in hw_model_list:
                                    hw_model = lines[indexes[m]+1]
                                    sh_ver_count += 1

                            indexes = [x for x, y in enumerate(lines) if y == "IOS"]   
                            for m in range(len(indexes)): 
                                if lines[indexes[m]+1] == "(tm)":
                                    sw_ver = lines[indexes[m]+6] 
                                    sh_ver_count += 1

                            indexes = [x for x, y in enumerate(lines) if y == "uptime"]   
                            for m in range(len(indexes)): 
                                if lines[indexes[m]+1] == "is":
                                    for l in range(5):
                                        if lines[indexes[m]+2+l] == "System": #stop here
                                            break
                                        else:
                                            uptime_list.append(lines[indexes[m]+2+l])
                                    uptime = ' '.join(uptime_list)   
                                    sh_ver_count += 1

                            if sh_ver_count != 0:
                                write_list1 = [hostname,"1.system","1.hardware model","","show ver",hw_model]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                write_list2 = [hostname,"1.system","2.software version","","show ver",sw_ver]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list2[j])
                                write_list3 = [hostname,"1.system","3.uptime","","show ver",uptime]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list3[j])
                                    
                            else:
                                write_list1 = [hostname,"1.system",""]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill  
                                write_list3 = ["",'Check if "show ver" is included in the log file.']
                                for j in range(len(write_list3)):
                                    ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
 
#####
#####show ip int brief/show cdp nei
#####

#############################sub-title
                            row_number2 += 1
                            title_row2 = ["hostname","2.interface","1.ip","flag","command","interface","status","IP address","subnet","description","CDP"]
                            for s in range(0, 11):
                                ws.cell(row=row_number2, column=s+1, value=title_row2[s])
                            for g in range(len(column_list)):
                                ws[column_list[g]+str(row_number2)].fill = blue_fill


#############################Interface & IP Address & Subnet Mask
                            sh_ip_int_count = 0
                            indexes = [x for x, y in enumerate(lines) if y == "interface"]   
                            for m in range(len(indexes)):  
                                interface = ""#initialize                   
                                ip_address = ""#initialize
                                subnet_mask = ""#initialize
                                description = ""#initialize
                                description_list = []#initialize  
                                cdp_nei = ""   #initialize                              
                                if re.match("ATM" + r'\d+/\d+',lines[indexes[m]+1].strip('\n')):
                                    continue 
                                    sh_ip_int_count +=1                                    
                                elif re.match("Loopback" + r'\d+', lines[indexes[m]+1].strip('\n')) or re.match("FastEthernet" + r'\d+/\d+', lines[indexes[m]+1].strip('\n')) or re.match("GigabitEthernet" + r'\d+/\d+', lines[indexes[m]+1].strip('\n')):
                                    interface = lines[indexes[m]+1]
                                    sh_ip_int_count +=1

                                    for k in range(0,10):
                                        if lines[indexes[m]+1+k] == "interface":#next interface
                                            break
                                        elif lines[indexes[m]+1+k] == "ip" and re.match(r'\d+.\d+.\d+.\d+',lines[indexes[m]+1+k+2]):
                                            ip_address = lines[indexes[m]+1+k+2]                               
                                            subnet_mask = lines[indexes[m]+1+k+3]                                    
                                        elif lines[indexes[m]+1+k] == "description":   
                                            for p in range(1,10):
                                                if lines[indexes[m]+1+k+p] == "ip" and lines[indexes[m]+1+k+p+1] == "address":
                                                    break
                                                else:
                                                    description_list.append(lines[indexes[m]+1+k+p])
                                            description = ' '.join(description_list)                                  

#############################Status                              
                                status = ""#initialize
                                indexes2 = [z for z, w in enumerate(lines) if w == interface.strip('\n')]
                                red_flag = "off"
                                for l in range(len(indexes2)):                                
                                    if lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "manual" or lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "NVRAM" or lines[indexes2[l]+2] == "YES" and lines[indexes2[l]+3] == "unset":
                                        if lines[indexes2[l]+5] == "up":
                                            status = "up"
                                            red_flag ="off"
                                        elif lines[indexes2[l]+4] == "administratively":  
                                            status = "admin down"
                                            red_flag ="gray"
                                        else:
                                            status = "down"  
                                            red_flag ="on"  
#############################CDP
                                cdp_int_list = []
                                if re.match("FastEthernet" + r'\d+/\d+', interface) or re.match("GigabitEthernet" + r'\d+/\d+', interface):#not loopback
                                    cdp_int = interface.replace('FastEthernet','Fas ')
                                    cdp_int = cdp_int.replace('GigabitEthernet','Gig ')
                                    cdp_int_list = cdp_int.split(" ")
                                    cdp_int = cdp_int_list[0] #such as "Fas"
                                    cdp_int_number = cdp_int_list[1].strip('\n') #such as "0/0"
                                    indexes2 = [x for x, y in enumerate(lines) if y == cdp_int]   
                                    for k in range(len(indexes2)): 
                                        if lines[indexes2[k]+3] =="R" and lines[indexes2[k]+1] == cdp_int_number: 
                                            cdp_nei = lines[indexes2[k]-1] + " " + lines[indexes2[k]+5] + lines[indexes2[k]+6]
                                        else:
                                            cdp_nei == ""

#############################write
                                if sh_ip_int_count != 0:
                                    if red_flag == "off": #up
                                        write_list1 = [hostname,"2.interface","1.ip","","show ip int brief/show cdp nei",interface]
                                        row_number2 += 1
                                        for j in range(len(write_list1)):
                                            ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                        write_list2 = [status]
                                        for j in range(len(write_list2)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                            ws['G'+str(row_number2)].fill = green_fill  
                                        write_list3 = [ip_address, subnet_mask, description, cdp_nei]
                                        for j in range(len(write_list3)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                                    elif red_flag == "gray": #admin down
                                        write_list1 = [hostname,"2.interface","1.ip","","show ip int brief/show cdp nei",interface]
                                        row_number2 += 1
                                        for j in range(len(write_list1)):
                                            ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                        write_list2 = [status]
                                        for j in range(len(write_list2)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                            ws['G'+str(row_number2)].fill = gray_fill  
                                        write_list3 = [ip_address, subnet_mask, description]
                                        for j in range(len(write_list3)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])                                             
                                    elif red_flag == "on": #down
                                        write_list1 = [hostname,"2.interface","1.ip","","show ip int brief/show cdp nei",interface]
                                        row_number2 += 1
                                        for j in range(len(write_list1)):
                                            ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                        write_list2 = [status]
                                        for j in range(len(write_list2)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                            ws['G'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                            ws['G'+str(row_number2)].fill = red_fill  
                                        write_list3 = [ip_address, subnet_mask, description]
                                        for j in range(len(write_list3)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j]) 
                            if sh_ip_int_count == 0:
                                write_list1 = [hostname,"2.interface","1.ip"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill  
                                write_list3 = ["",'Check if "show ip int brief/show cdp nei" is included in the log file.']
                                for j in range(len(write_list3)): 
                                    ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])




#####
#####show ip ospf nei
#####
                            router_ospf_count = 0
                            no_passive_list = []
                            no_passive_count = 0
                            ospf_nei_count = 0
                            ospf_nei = "" #initialize
                            ospf_nei_state = "" #initialize
                            ospf_indexes = [x for x, y in enumerate(lines) if y == "ospf"]
                            for q in range(len(ospf_indexes)):
                                if lines[ospf_indexes[q]-1] == "router":
                                    router_ospf_count += 1
                                    for p in range(30):
                                        if lines[ospf_indexes[q]+p] == "!": #stop here
                                            break
                                        elif lines[ospf_indexes[q]+p] == "no" and lines[ospf_indexes[q]+p+1] == "passive-interface":
                                            no_passive_list.append(lines[ospf_indexes[q]+p+2])
                                            no_passive_count += 1
                            for r in range(len(no_passive_list)):
                                no_passive_indexes = [x for x, y in enumerate(lines) if y == no_passive_list[r]]
                                for s in range(len(no_passive_indexes)):
                                    if re.match(r'\d+.\d+.\d+.\d+',lines[no_passive_indexes[s]-1]):
                                        ospf_int = lines[no_passive_indexes[s]]
                                        ospf_nei_state = lines[no_passive_indexes[s]-4]
                                        ospf_nei = lines[no_passive_indexes[s]-6]
                                        ospf_nei_count += 1
                                        if re.search("FULL" +r'\S+',ospf_nei_state):#FULL
                                            write_list1 = [hostname,"3.routing","1.ospf","","show run/show ip ospf nei",ospf_int,]
                                            row_number2 += 1
                                            for j in range(len(write_list1)):
                                                ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                            write_list2 = [ospf_nei_state]
                                            for j in range(len(write_list2)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                                ws['G'+str(row_number2)].fill = green_fill
                                            write_list3 = [ospf_nei]
                                            for j in range(len(write_list3)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j]) 
                                        else:#not FULL
                                            write_list1 = [hostname,"3.routing","1.ospf"]
                                            row_number2 += 1
                                            for j in range(len(write_list1)):
                                                ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                            write_list2 = ["error!!"]
                                            for j in range(len(write_list2)):                                            
                                                ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                                ws['D'+str(row_number2)].fill = red_fill
                                                ws['D'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                            write_list3 = ["show run/show ip ospf nei",ospf_int,]
                                            for j in range(len(write_list3)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                                            write_list4 = [ospf_nei_state]
                                            for j in range(len(write_list4)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2)+len(write_list3), value=write_list4[j])
                                                ws['G'+str(row_number2)].fill = red_fill
                                                ws['G'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                            write_list5 = [ospf_nei]
                                            for j in range(len(write_list5)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2)+len(write_list3)+len(write_list4), value=write_list5[j])                                  
                            
                            if router_ospf_count == 0:
                                write_list1 = [hostname,"3.routing","1.ospf"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):                                            
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill
                                    write_list3 = ["show run/show ip ospf nei",'This device may not support OSPF or "show run" is missing']
                                    for j in range(len(write_list3)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                            
                            else:
                                if no_passive_count != ospf_nei_count:
                                    write_list1 = [hostname,"3.routing","1.ospf"]
                                    row_number2 += 1
                                    for j in range(len(write_list1)):
                                        ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                    write_list2 = ["check!!"]
                                    for j in range(len(write_list2)):                                            
                                        ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                        ws['D'+str(row_number2)].fill = yellow_fill
                                        write_list3 = ["show run/show ip ospf nei",'Check if any OSPF neighbor is missing. No-passive count:' + str(no_passive_count) + " Neighbor count:"+ str(ospf_nei_count) +"."]
                                        for j in range(len(write_list3)):
                                            ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                            
#####
#####show ip bgp summary
#####                                             

                            router_bgp_count = 0
                            bgp_nei_list = []
                            bgp_nei_count = 0
                            my_as = "" #initialize
                            remote_as = "" #initialize
                            prefix_no = "" #initialize
                            bgp_state_list = ["Idle", "Active", "Connect", "OpenSent","OpenConfirm"]
                            bgp_indexes = [x for x, y in enumerate(lines) if y == "bgp"]
                            for q in range(len(bgp_indexes)):
                                if lines[bgp_indexes[q]-1] == "router":
                                    router_bgp_count += 1
                                    my_as = lines[bgp_indexes[q]+1]
                                    for p in range(40):
                                        if lines[bgp_indexes[q]+p] == "!": #stop here
                                            break
                                        elif lines[bgp_indexes[q]+p] == "neighbor" and lines[bgp_indexes[q]+p+2] == "remote-as" or lines[bgp_indexes[q]+p] == "neighbor" and lines[bgp_indexes[q]+p+2] == "peer-group":
                                            bgp_nei_list.append(lines[bgp_indexes[q]+p+1])
                                            bgp_nei_count += 1
                            for r in range(len(bgp_nei_list)):
                                bgp_nei_indexes = [x for x, y in enumerate(lines) if y == bgp_nei_list[r]]
                                for s in range(len(bgp_nei_indexes)):
                                    if lines[bgp_nei_indexes[s]+1] == "4": #version No.
                                        bgp_nei_address = lines[bgp_nei_indexes[s]]
                                        remote_as = lines[bgp_nei_indexes[s]+2]
                                        prefix_no = lines[bgp_nei_indexes[s]+9]
                                        if prefix_no.strip('\n') not in bgp_state_list:
                                            write_list1 = [hostname,"3.routing","2.bgp","","show run/show ip bgp summary","local:"+my_as]
                                            row_number2 += 1
                                            for j in range(len(write_list1)):
                                                ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                            write_list2 = ["Established:"+prefix_no]
                                            for j in range(len(write_list2)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                                ws['G'+str(row_number2)].fill = green_fill
                                            write_list3 = [bgp_nei_address,"remote:"+remote_as]
                                            for j in range(len(write_list3)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j]) 
                                        else:
                                            write_list1 = [hostname,"3.routing","2.bgp"]
                                            row_number2 += 1
                                            for j in range(len(write_list1)):
                                                ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                            write_list2 = ["error!!"]
                                            for j in range(len(write_list2)):                                            
                                                ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                                ws['D'+str(row_number2)].fill = red_fill
                                                ws['D'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                            write_list3 = ["show run/show ip bgp summary","local:"+my_as]
                                            for j in range(len(write_list3)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                                            write_list4 = [prefix_no]
                                            for j in range(len(write_list4)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2)+len(write_list3), value=write_list4[j])
                                                ws['G'+str(row_number2)].fill = red_fill
                                                ws['G'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                            write_list5 = [bgp_nei_address,"remote:"+remote_as]
                                            for j in range(len(write_list5)):
                                                ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2)+len(write_list3)+len(write_list4), value=write_list5[j])


                            if router_bgp_count == 0:
                                write_list1 = [hostname,"3.routing","2.bgp"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):                                            
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill
                                    write_list3 = ["show run/show ip bgp summary",'This device may not support BGP or "show run" is missing']
                                    for j in range(len(write_list3)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
   
#####
#####show ip route
#####  

                            default_route_count = 0
                            default_route_list = []
                            route_flag_list = ["O", "O*E2", "O*E1", "C", "L", "B", "B*", "S", "D"]
                            default_route_indexes = [x for x, y in enumerate(lines) if y == "0.0.0.0/0"]
                            for s in range(len(default_route_indexes)):
                                if lines[default_route_indexes[s]+2] == "via" and lines[default_route_indexes[s]+7] == "via": # with interface
                                    default_route_count += 1
                                    if lines[default_route_indexes[s]-2] == "O" or lines[default_route_indexes[s]-2] == "B": 
                                        default_route_list.append(lines[default_route_indexes[s]-2])
                                        default_route_list.append(lines[default_route_indexes[s]-1])
                                        for t in range(11):
                                            default_route_list.append(lines[default_route_indexes[s]+t])
                                    elif lines[default_route_indexes[s]-1] == "O" or lines[default_route_indexes[s]-1] == "O*E1" or lines[default_route_indexes[s]-1] == "O*E2" or lines[default_route_indexes[s]-1] == "O*IA" or lines[default_route_indexes[s]-1] == "B" or lines[default_route_indexes[s]-1] == "B*":
                                        default_route_list.append(lines[default_route_indexes[s]-1])
                                        for t in range(11):
                                            default_route_list.append(lines[default_route_indexes[s]+t])
                                elif lines[default_route_indexes[s]+2] == "via" and lines[default_route_indexes[s]+6] == "via": # without interface)
                                    default_route_count += 1
                                    if lines[default_route_indexes[s]-2] == "O" or lines[default_route_indexes[s]-2] == "B": 
                                        default_route_list.append(lines[default_route_indexes[s]-2])
                                        default_route_list.append(lines[default_route_indexes[s]-1])
                                        for t in range(9):
                                            default_route_list.append(lines[default_route_indexes[s]+t])                                         
                                    elif lines[default_route_indexes[s]-1] == "O" or lines[default_route_indexes[s]-1] == "O*E1" or lines[default_route_indexes[s]-1] == "O*E2" or lines[default_route_indexes[s]-1] == "O*IA" or lines[default_route_indexes[s]-1] == "B" or lines[default_route_indexes[s]-1] == "B*":
                                        default_route_list.append(lines[default_route_indexes[s]-1])
                                        for t in range(9):
                                            default_route_list.append(lines[default_route_indexes[s]+t])
                                                                          
                            if len(default_route_list) != 0:
                                write_list1 = [hostname,"3.routing","3.default route","","show ip route"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                default_route_string = ' '.join(default_route_list)
                                default_route_string = default_route_string.replace("\n","\r\n")
                                ws.cell(row=row_number2, column=1+len(write_list1), value= default_route_string)
                                ws['F'+str(row_number2)].alignment = Alignment(wrapText = True)   #wrap text                             
                            if default_route_count == 0:
                                write_list1 = [hostname,"3.routing","3.default route"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):                                            
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill
                                    write_list3 = ["show ip route",'This device may not have default routes or "show ip route" is missing']
                                    for j in range(len(write_list3)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])

#####
#####show ntp associations
#####  

                            ntp_count = 0
                            ntp_indexes = [x for x, y in enumerate(lines) if re.match("\*~" + r'\d+.\d+.\d+.\d+' ,y) or re.match("\+~" + r'\d+.\d+.\d+.\d+' ,y) or re.match("#~" + r'\d+.\d+.\d+.\d+' ,y) or re.match("\-~" + r'\d+.\d+.\d+.\d+' ,y) or re.match("~" + r'\d+.\d+.\d+.\d+' ,y)]
                            for s in range(len(ntp_indexes)):
                                ntp_status = lines[ntp_indexes[s]]
                                if ntp_status[0:2] == r"*~" or ntp_status[0:2] == r"+~":
                                    ntp_count += 1                                    
                                    write_list1 = [hostname,"4.ntp","1.ntp status","","show ntp associations",""]
                                    row_number2 += 1
                                    for j in range(len(write_list1)):
                                        ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                    write_list2 = [ntp_status]
                                    for j in range(len(write_list2)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                        ws['G'+str(row_number2)].fill = green_fill
                                else:
                                    ntp_count += 1  
                                    write_list1 = [hostname,"4.ntp","1.ntp status"]
                                    row_number2 += 1
                                    for j in range(len(write_list1)):
                                        ws.cell(row=row_number2, column=j+1, value=write_list1[j])
                                    write_list2 = ["error!!"]
                                    for j in range(len(write_list2)):                                            
                                        ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                        ws['D'+str(row_number2)].fill = red_fill
                                        ws['D'+str(row_number2)].font = Font(color='FFFFFF') #white characters
                                    write_list3 = ["show ntp associations",""]
                                    for j in range(len(write_list3)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])
                                    write_list4 = [ntp_status]
                                    for j in range(len(write_list4)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2)+len(write_list3), value=write_list4[j])
                                        ws['G'+str(row_number2)].fill = red_fill
                                        ws['G'+str(row_number2)].font = Font(color='FFFFFF') #white characters

                            if ntp_count != 2:
                                write_list1 = [hostname,"4.ntp","1.ntp status"]
                                row_number2 += 1
                                for j in range(len(write_list1)):
                                    ws.cell(row=row_number2, column=j+1, value=write_list1[j])  
                                write_list2 = ["check!!"]
                                for j in range(len(write_list2)):                                            
                                    ws.cell(row=row_number2, column=j+1+len(write_list1), value=write_list2[j])
                                    ws['D'+str(row_number2)].fill = yellow_fill
                                    write_list3 = ["show ntp associations",'Check if NTP is configured or "show ntp associaitons" is included in the log file.']
                                    for j in range(len(write_list3)):
                                        ws.cell(row=row_number2, column=j+1+len(write_list1)+len(write_list2), value=write_list3[j])

#####
#####save
#####


                            side1 = Side(style='thin', color='000000')
                            border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
 
                            for row in ws['A1:K'+str(row_number2)]:
                                for cell in row:
                                    cell.border = border_aro     #line
                            wb.save(result_file) 
                except IndexError:
                    pass



        os.popen(result_file)
        




  

if __name__ == '__main__':
    #current_directory = os.getcwd()
    current_directory = os.path.dirname(os.path.abspath(__file__))
    initial_screen()
